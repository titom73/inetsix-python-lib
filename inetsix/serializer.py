import logging
import openpyxl
from openpyxl.utils.exceptions import (
    InvalidFileException,
    SheetTitleException
)


class ExcelSerializer:
    """ Convert Excel data into JSON structures.

    Examples:
    ---------
    >>> serializer = ExcelSerializer(excel_path='bin/Book1.xlsx')
    >>> json_result = serializer.serialize_table(sheet='Sheet1')
    >>> pp.pprint(json_result)
        [ {'description': 'ifd1', 'interface': 'ethernet1', 'state': 'up'},
          {'description': 'ifd2', 'interface': 'ethernet2', 'state': 'down'}]
    >>>
    >>> json_result = serializer.serialize_book()
    >>> pp.pprint(json_result)
        { 'Sheet1':
            [
              {'description': 'ifd1', 'interface': 'ethernet1', 'state': 'up'},
              { 'description': 'ifd2', 'interface': 'ethernet2', 'state': 'down'}
            ]
        }
    """

    def __init__(self, excel_path):
        """
        Class Constructor

        Parameters
        ----------
        excel_path : [type], optional
            Path to excel file you want to read data.
        """
        self._logger = logging.getLogger('inetsix.serializer.ExcelSerializer')
        self._logger.info('Excel serializer initiated')
        self._filename = excel_path
        try:
            self._book = openpyxl.load_workbook(self._filename, data_only=True)
        except InvalidFileException as exception_message:
            self._logger.error('Error: %s', str(exception_message))
            raise

    def serialize_table(self, sheet=None):
        """
        Generate JSON/list structure for given sheet.

        Parameters
        ----------
        sheet : string
            Name of the sheet to serialize.

        Returns
        -------
        list
            serialized content of all sheets.
        """
        self._logger.info('Serializing sheet %s', str(sheet))
        try:
            active = self._book.get_sheet_by_name(sheet)
        except SheetTitleException as exception_message:
            self._logger.error('Error: %s', str(exception_message))
        res = list(active)
        final = list()
        for x in range(1, len(res)):
            partFinal = {}
            partFinal[res[0][0].value] = res[x][0].value
            partFinal[res[0][1].value] = res[x][1].value
            partFinal[res[0][2].value] = res[x][2].value
            final.append(partFinal)
        return final

    @property
    def get_sheets_name(self):
        """
        Returns list of sheets found in Excel Workbook.

        Returns
        -------
        list
            list of sheets
        """
        self._logger.info('list all sheets in book %s', self._filename)
        self._logger.debug('sheets found: %s', str(self._book.sheetnames))
        return self._book.sheetnames

    def serialize_book(self):
        """
        Generate JSON/dict structure for all sheets available.

        Returns
        -------
        dict
            serialized content of all sheets.
        """
        self._logger.info('Serialize all sheets in book %s', self._filename)
        json_book = dict()
        for sheet_name in self.get_sheets_name:
            self._logger.debug('  - working on %s', sheet_name)
            json_book[sheet_name] = self.serialize_table(sheet=sheet_name)
        return json_book
